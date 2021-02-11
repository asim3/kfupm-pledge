from django.db.models import Model, FileField
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from django.contrib.auth.models import User
from openpyxl import load_workbook
from student.models import Pledge
from threading import Thread


def extract_excel_file(excel_file_path):
    excel_file = load_workbook(excel_file_path)
    for sheet in excel_file.sheetnames:
        for (i, row) in enumerate(excel_file[sheet].rows):
            if i < 1:
                continue
            add_new_student(row, sheet.capitalize())


def add_new_student(excel_row, pledge_type):
    kfupmid = str(excel_row[1].value)
    try:
        user = User.objects.get(username=kfupmid)
    except User.DoesNotExist:
        user = User.objects.create_user(
            kfupmid,
            email=f's{kfupmid}@kfupm.edu.sa',
            password=str(excel_row[2].value),
            first_name=excel_row[3].value)

    next_tearm = excel_row[4].value
    kfupm_gpa = excel_row[5].value if len(excel_row) > 5 else 0.0
    try:
        Pledge.objects.get(
            student=user,
            pledge_type=pledge_type,
            next_tearm=next_tearm)
        print('get')
    except Pledge.DoesNotExist:
        Pledge.objects.create(
            student=user,
            pledge_type=pledge_type,
            kfupm_gpa=kfupm_gpa,
            next_tearm=next_tearm)
        print('add')


class AddStudents(Model):

    class Meta:
        ordering = ["excel_file"]
        verbose_name = _('Add Student')
        verbose_name_plural = _("Add Students")

    excel_file = FileField(upload_to='uploads/%Y/%m/%d/')

    def get_absolute_url(self):
        return reverse_lazy('admin-student')

    def save(self, *args, **kwargs):
        super().save(*args, **kwargs)
        Thread(target=extract_excel_file, args=(self.excel_file.path,)).start()
