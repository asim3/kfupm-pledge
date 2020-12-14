from django.views.generic import CreateView
from django.contrib.auth.models import User
from openpyxl import load_workbook
from .forms import AddStudentsForm
from student.models import Pledge


def add_new_student(row, pledge_type):
    kfupmid = row[1].value
    passwd = str(row[2].value)
    name = row[3].value
    tearm = row[4].value
    gpa = row[5].value if len(row) > 5 else 0.0
    user = User.objects.create_user(kfupmid, f's{kfupmid}@kfupm.edu.sa',
                                    passwd, first_name=name)
    Pledge.objects.create(
        student=user,
        pledge_type=pledge_type,
        kfupm_gpa=gpa,
        next_tearm=tearm)


def add_new_students_from_excel(e_path):
    e_file = load_workbook(e_path)
    for sheet in e_file.sheetnames:
        for (i, row) in enumerate(e_file[sheet].rows):
            if i < 1:
                continue
            add_new_student(row, sheet.capitalize())


class StudentsView(CreateView):
    """
    Adding students using an excel file.

    - List all student registered in the database, 
    - download the template excel file.
    """
    template_name = "administration/student.html"
    form_class = AddStudentsForm
    success_url = "/admin/student/pledge/"

    def form_valid(self, form):
        response = super().form_valid(form)
        add_new_students_from_excel(self.object.excel_file.path)
        return response
