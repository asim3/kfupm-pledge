from django.forms import ModelForm, ValidationError
from django.utils.translation import gettext_lazy as _
from openpyxl import load_workbook
from zipfile import BadZipFile

from student.models import Pledge
from .models import AddStudents


class AddStudentsForm(ModelForm):
    class Meta:
        model = AddStudents
        fields = ['excel_file']

    def clean_excel_file(self):
        excel_file = self.cleaned_data.get('excel_file')
        try:
            e_file = load_workbook(excel_file.file)
        except BadZipFile as e:
            raise ValidationError(_("Excel file is not valid!"))

        for sheet in e_file.sheetnames:
            for (i, row) in enumerate(e_file[sheet].rows):
                if i < 1:
                    continue
                kfupmid = row[1].value
                next_tearm = row[4].value
                if isinstance(kfupmid, int):
                    if Pledge.objects.filter(student__username=kfupmid, next_tearm=next_tearm):
                        raise ValidationError(
                            _("Student {} pledge is already listed in {} Tearm").format(kfupmid, next_tearm))
                else:
                    raise ValidationError(
                        _("Row number {} in {} is not valid!").format(i, sheet))
                if not isinstance(row[2].value, int):
                    raise ValidationError(
                        _("Student {} national ID is not valid!").format(kfupmid))
                if not isinstance(next_tearm, int):
                    raise ValidationError(
                        _("Student {} next tearm is not valid!").format(kfupmid))
                if sheet != Pledge.PledgeType.DISMISS and not isinstance(row[5].value, float):
                    if not isinstance(row[5].value, int):
                        raise ValidationError(
                            _("Student {} GPA is not valid!").format(kfupmid))
        return excel_file
