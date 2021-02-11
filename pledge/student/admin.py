from django.contrib.admin import register, ModelAdmin
from django.contrib import messages
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from django.utils.timezone import datetime
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from io import BytesIO
from zipfile import ZipFile

from .models import Pledge
from .utils import template_to_pdf


@register(Pledge)
class AdminPledge(ModelAdmin):
    change_list_template = 'reports/tools.html'
    search_fields = ('student__username', 'student__first_name',
                     'phone', 'phone_guardian')
    list_filter = ('pledge_type', 'next_tearm', 'is_approved', 'export_date',)
    ordering = ('date_added',)
    list_display = (
        'student',
        'pledge_type',
        'next_tearm',
        'kfupm_gpa',
        'date_added',
        'phone',
        'phone_guardian',
        'approved_date',
        'export_date',
        'is_approved',)

    def export_as_excel(self, request, queryset):
        excel_file = Workbook()
        excel_file.active.title = Pledge.PledgeType.CONDITION
        excel_file.create_sheet(title=Pledge.PledgeType.FINAL)
        excel_file.create_sheet(title=Pledge.PledgeType.DISMISS)
        settings = (
            ('A', 5, _('No')),
            ('B', 15, _('KFUPM ID')),
            ('C', 30, _('Name')),
            ('D', 30, _('Approved Date')),
            ('E', 15, _('Phone')),
            ('F', 15, _('Phone Guardian')),
        )

        for sheet_name in excel_file.sheetnames:
            sheet = excel_file[sheet_name]
            for setting in settings:
                sheet.column_dimensions[setting[0]].width = setting[1]
                sheet['%s1' % setting[0]] = str(setting[2])

            data = queryset.filter(pledge_type=sheet_name)
            for i, pledge in enumerate(data, start=2):
                sheet['A%d' % i] = i - 1
                sheet['B%d' % i] = pledge.student.username
                sheet['C%d' % i] = pledge.student.first_name
                sheet['D%d' % i] = pledge.approved_date
                sheet['E%d' % i] = pledge.phone
                sheet['F%d' % i] = pledge.phone_guardian

        queryset.update(export_date=datetime.now())

        excel_bytes = save_virtual_workbook(excel_file)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = 'pledge-%s.xlsx' % datetime.now().strftime('%s')
        response = HttpResponse(excel_bytes, content_type=content_type)
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        return response

    def export_as_pdf(self, request, queryset):
        if 10 < queryset.count():
            self.message_user(
                request, 'العدد الاقصى المسموح هو 10', messages.WARNING)
        else:
            pledges = queryset.filter(is_approved=True)
            if not pledges:
                self.message_user(request, 'لم يتم التحديد', messages.WARNING)
                return None
            temporary_file = BytesIO()
            zip = ZipFile(temporary_file, "a")

            for pledge in pledges:
                pdf_file = template_to_pdf(
                    'pdf/pledge.html', {'object': pledge})
                zip.writestr("id-%s.pdf" %
                             (pledge.student.username,), pdf_file.read())

            # fix for Linux zip files read in Windows
            for file in zip.filelist:
                file.create_system = 0
            zip.close()

            filename = 'pledge-%s.zip' % datetime.now().strftime('%s')
            temporary_file.seek(0)
            response = HttpResponse(
                temporary_file.read(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="%s"' % filename
            temporary_file.close()
            return response

    export_as_excel.short_description = "تصدير إلى ملف إكسل"
    export_as_pdf.short_description = "تصدير إلى ملف PDF"
    actions = ['export_as_excel', 'export_as_pdf']
